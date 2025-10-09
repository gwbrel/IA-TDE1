import speech_recognition as sr
import openpyxl
import webbrowser
import os

# Cria ou abre a planilha
FILE = "lista_compras.xlsx"
if not os.path.exists(FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"
    ws.append(["Produto"])
    wb.save(FILE)

def adicionar_produto(produto):
    wb = openpyxl.load_workbook(FILE)
    ws = wb.active
    ws.append([produto])
    wb.save(FILE)
    print(f"{produto} adicionado à lista!")

def listar_produtos():
    wb = openpyxl.load_workbook(FILE)
    ws = wb.active
    produtos = [cell.value for cell in ws["A"][1:]]
    print("Lista de compras:", produtos)
    return produtos

def enviar_whatsapp():
    produtos = listar_produtos()
    mensagem = "Lista de compras: " + ", ".join(produtos)
    link = "https://wa.me/?text=" + mensagem.replace(" ", "%20")
    webbrowser.open(link)

def reconhecer_voz():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("Fale o comando...")
        audio = r.listen(source)
        try:
            comando = r.recognize_google(audio, language="pt-BR").lower()
            print("Você disse:", comando)
            return comando
        except:
            print("Não entendi, tente novamente.")
            return ""

def main():
    while True:
        comando = reconhecer_voz()
        if "adicionar" in comando:
            produto = comando.replace("adicionar", "").strip()
            adicionar_produto(produto)
        elif "listar" in comando:
            listar_produtos()
        elif "enviar" in comando:
            enviar_whatsapp()
        elif "sair" in comando:
            print("Encerrando agente...")
            break

if __name__ == "__main__":
    main()
